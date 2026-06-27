#!/usr/bin/env python3
"""
Final Clearance Team Monthly Scorecard Generator
Analyzes 6 Excel files including SLA tracking
"""

import openpyxl
from datetime import datetime, timedelta
import json
from pathlib import Path
from collections import defaultdict
import re

class FinalClearanceScorecard:
    def __init__(self, folder_path):
        self.folder = Path(folder_path)
        self.employees = defaultdict(dict)
        self.current_month = datetime.now().month
        self.final_clearance_team = [
            'Archana Gautam',
            'Aswani R',
            'Anubha Priyam'
        ]
        self.sla_target_days = 2  # SLA target: 2 days
        
    def normalize_name(self, name):
        """Normalize employee names"""
        if not name:
            return ""
        name = str(name).strip().lower()
        name = ' '.join(name.split())
        return name
    
    def is_final_clearance_team(self, name):
        """Check if employee is in Final Clearance Team"""
        if not name:
            return False
        name_norm = self.normalize_name(name)
        for team_member in self.final_clearance_team:
            if self.normalize_name(team_member) in name_norm or name_norm in self.normalize_name(team_member):
                return True
        return False
    
    def load_production(self):
        """Load Production Tracker"""
        print("\n⚡ Loading Production Tracker...")
        file_path = self.folder / "Production_Tracker.xlsx"
        
        if not file_path.exists():
            print(f"   ❌ File not found")
            return
        
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        
        for row_idx in range(2, ws.max_row + 1):
            emp_name = ws.cell(row=row_idx, column=1).value
            if not emp_name or not self.is_final_clearance_team(emp_name):
                continue
            
            emp_name = str(emp_name).strip()
            hours = ws.cell(row=row_idx, column=self.current_month + 1).value
            
            try:
                hours = float(hours) if hours else 0
            except:
                hours = 0
            
            status = 'Green' if hours >= 8 else 'Red'
            
            self.employees[emp_name]['productivity_hours'] = hours
            self.employees[emp_name]['productivity_status'] = status
        
        wb.close()
        print(f"   ✅ Loaded production data")
    
    def load_attendance(self):
        """Load Attendance"""
        print("\n📊 Loading Attendance...")
        file_path = self.folder / "Attendance.xlsx"
        
        if not file_path.exists():
            print(f"   ❌ File not found")
            return
        
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        
        for row_idx in range(2, ws.max_row + 1):
            emp_name = ws.cell(row=row_idx, column=1).value
            if not emp_name or not self.is_final_clearance_team(emp_name):
                continue
            
            emp_name = str(emp_name).strip()
            leaves = ws.cell(row=row_idx, column=self.current_month + 1).value
            
            try:
                leaves = int(leaves) if leaves else 0
            except:
                leaves = 0
            
            working_days = 22
            attendance_pct = max(0, ((working_days - leaves) / working_days) * 100)
            
            self.employees[emp_name]['attendance_leaves'] = leaves
            self.employees[emp_name]['attendance_pct'] = round(attendance_pct, 2)
        
        wb.close()
        print(f"   ✅ Loaded attendance data")
    
    def load_pkt(self):
        """Load Process Knowledge Test"""
        print("\n📚 Loading PKT...")
        file_path = self.folder / "Process_Knowledge_Test.xlsx"
        
        if not file_path.exists():
            print(f"   ❌ File not found")
            return
        
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        
        for row_idx in range(2, ws.max_row + 1):
            emp_name = ws.cell(row=row_idx, column=1).value
            if not emp_name or not self.is_final_clearance_team(emp_name):
                continue
            
            emp_name = str(emp_name).strip()
            score = ws.cell(row=row_idx, column=self.current_month + 1).value
            
            try:
                score = int(score) if score else 0
            except:
                score = 0
            
            if score >= 90:
                rating = 'Excellent'
            elif score >= 80:
                rating = 'Good'
            else:
                rating = 'Needs Improvement'
            
            self.employees[emp_name]['pkt_score'] = score
            self.employees[emp_name]['pkt_rating'] = rating
        
        wb.close()
        print(f"   ✅ Loaded PKT data")
    
    def load_internal_audit(self):
        """Load Internal Audit Errors"""
        print("\n🔍 Loading Internal Audit...")
        file_path = self.folder / "Internal_Audit_Scores.xlsx"
        
        if not file_path.exists():
            print(f"   ❌ File not found")
            return
        
        wb = openpyxl.load_workbook(file_path)
        
        if 'Internal Audit' in wb.sheetnames:
            ws = wb['Internal Audit']
            error_count = defaultdict(int)
            
            for row_idx in range(2, ws.max_row + 1):
                team_member_cell = ws.cell(row=row_idx, column=5).value
                error_cell = ws.cell(row=row_idx, column=2).value
                
                if team_member_cell and error_cell:
                    try:
                        errors = int(str(error_cell).split()[0])
                        members_str = str(team_member_cell)
                        for team_member in self.final_clearance_team:
                            if self.normalize_name(team_member) in members_str.lower():
                                error_count[team_member] += errors
                    except:
                        pass
        
        for emp_name in error_count:
            errors = error_count[emp_name]
            if errors == 0:
                status = 'Green'
            elif errors <= 2:
                status = 'Amber'
            else:
                status = 'Red'
            
            self.employees[emp_name]['audit_errors'] = errors
            self.employees[emp_name]['audit_status'] = status
        
        wb.close()
        print(f"   ✅ Loaded internal audit data")
    
    def load_final_clearance(self):
        """Load Final Clearance Tracker with SLA"""
        print("\n📋 Loading Final Clearance Tracker...")
        file_path = self.folder / "Final_Clearance_Tracker.xlsx"
        
        if not file_path.exists():
            print(f"   ❌ File not found")
            return
        
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        
        fc_data = defaultdict(lambda: {
            'completed': 0,
            'pending': 0,
            'sla_days': [],
            'sla_met': 0,
            'sla_breached': 0
        })
        
        for row_idx in range(2, min(ws.max_row + 1, 500)):
            audited_by = ws.cell(row=row_idx, column=9).value
            date_received = ws.cell(row=row_idx, column=1).value
            fully_cleared_date = ws.cell(row=row_idx, column=14).value
            audit_status = ws.cell(row=row_idx, column=10).value
            
            if not audited_by or not self.is_final_clearance_team(audited_by):
                continue
            
            audited_by = str(audited_by).strip()
            
            # Calculate SLA
            if date_received and fully_cleared_date:
                try:
                    received_dt = date_received if isinstance(date_received, datetime) else datetime.strptime(str(date_received), '%Y-%m-%d')
                    cleared_dt = fully_cleared_date if isinstance(fully_cleared_date, datetime) else datetime.strptime(str(fully_cleared_date), '%Y-%m-%d')
                    sla_days = (cleared_dt - received_dt).days
                    
                    fc_data[audited_by]['sla_days'].append(sla_days)
                    fc_data[audited_by]['completed'] += 1
                    
                    if sla_days <= self.sla_target_days:
                        fc_data[audited_by]['sla_met'] += 1
                    else:
                        fc_data[audited_by]['sla_breached'] += 1
                except:
                    pass
            elif audit_status and 'Pending' in str(audit_status):
                fc_data[audited_by]['pending'] += 1
        
        for emp_name in fc_data:
            data = fc_data[emp_name]
            completed = data['completed']
            pending = data['pending']
            sla_days_list = data['sla_days']
            sla_met = data['sla_met']
            
            avg_sla = sum(sla_days_list) / len(sla_days_list) if sla_days_list else 0
            sla_compliance = (sla_met / completed * 100) if completed > 0 else 0
            
            self.employees[emp_name]['fc_completed'] = completed
            self.employees[emp_name]['fc_pending'] = pending
            self.employees[emp_name]['fc_avg_sla'] = round(avg_sla, 1)
            self.employees[emp_name]['fc_sla_compliance'] = round(sla_compliance, 1)
            self.employees[emp_name]['fc_status'] = 'Green' if sla_compliance >= 80 else 'Red' if sla_compliance < 50 else 'Amber'
        
        wb.close()
        print(f"   ✅ Loaded final clearance data")
    
    def load_client_audit(self):
        """Load Client System Audit"""
        print("\n🏢 Loading Client System Audit...")
        file_path = self.folder / "Client_System_Audit_Tracker.xlsx"
        
        if not file_path.exists():
            print(f"   ❌ File not found")
            return
        
        wb = openpyxl.load_workbook(file_path)
        ws = wb['Client System Audit']
        
        audit_pending = defaultdict(int)
        
        for row_idx in range(2, ws.max_row + 1):
            team_member = ws.cell(row=row_idx, column=3).value
            pending = ws.cell(row=row_idx, column=8).value
            
            if team_member and self.is_final_clearance_team(team_member):
                team_member = str(team_member).strip()
                try:
                    pending_count = int(pending) if pending else 0
                    audit_pending[team_member] += pending_count
                except:
                    pass
        
        for emp_name in audit_pending:
            count = audit_pending[emp_name]
            self.employees[emp_name]['client_audit_pending'] = count
            self.employees[emp_name]['client_audit_status'] = 'Red' if count > 0 else 'Green'
        
        wb.close()
        print(f"   ✅ Loaded client audit data")
    
    def calculate_overall_score(self):
        """Calculate Overall Score"""
        print("\n🎯 Calculating Overall Scores...")
        
        for emp_name in self.employees:
            emp = self.employees[emp_name]
            
            scores = []
            weights = []
            
            # PKT: 35%
            if 'pkt_score' in emp:
                scores.append(emp['pkt_score'])
                weights.append(35)
            
            # Productivity: 25%
            if 'productivity_hours' in emp:
                prod_score = min((emp['productivity_hours'] / 8) * 100, 100)
                scores.append(prod_score)
                weights.append(25)
            
            # SLA: 25%
            if 'fc_sla_compliance' in emp:
                scores.append(emp['fc_sla_compliance'])
                weights.append(25)
            
            # Attendance: 15%
            if 'attendance_pct' in emp:
                scores.append(emp['attendance_pct'])
                weights.append(15)
            
            if scores:
                total_weight = sum(weights)
                overall = sum(s * w for s, w in zip(scores, weights)) / total_weight if total_weight > 0 else 0
                emp['overall_score'] = round(overall, 2)
                
                if overall >= 90:
                    emp['overall_rating'] = 'Excellent'
                elif overall >= 80:
                    emp['overall_rating'] = 'Good'
                else:
                    emp['overall_rating'] = 'Needs Improvement'
                
                # Incentive eligibility
                audit_pending = emp.get('client_audit_pending', 0)
                if audit_pending > 0:
                    emp['incentive_eligible'] = 'Not Eligible'
                else:
                    emp['incentive_eligible'] = 'Eligible'
    
    def generate_scorecard(self):
        """Generate consolidated scorecard"""
        print("\n📊 Generating Final Clearance Team Scorecard...")
        
        scorecard = []
        for emp_name in sorted(self.employees.keys()):
            emp = self.employees[emp_name]
            
            record = {
                'Employee Name': emp_name,
                'Productivity Hours': emp.get('productivity_hours', '-'),
                'Productivity Status': emp.get('productivity_status', '-'),
                'Audit Errors': emp.get('audit_errors', 0),
                'Audit Status': emp.get('audit_status', '-'),
                'FC Completed': emp.get('fc_completed', 0),
                'FC Pending': emp.get('fc_pending', 0),
                'SLA Compliance %': emp.get('fc_sla_compliance', '-'),
                'Avg SLA Days': emp.get('fc_avg_sla', '-'),
                'PKT Score': emp.get('pkt_score', '-'),
                'PKT Rating': emp.get('pkt_rating', '-'),
                'Attendance Leaves': emp.get('attendance_leaves', '-'),
                'Attendance %': emp.get('attendance_pct', '-'),
                'Client Audit Pending': emp.get('client_audit_pending', 0),
                'Client Audit Status': emp.get('client_audit_status', '-'),
                'Overall Score': emp.get('overall_score', '-'),
                'Overall Rating': emp.get('overall_rating', '-'),
                'Incentive Eligible': emp.get('incentive_eligible', '-'),
                'Remarks': 'On Track' if emp.get('incentive_eligible') == 'Eligible' else 'Action Required'
            }
            scorecard.append(record)
        
        return scorecard
    
    def run(self):
        """Execute all data loading"""
        print("\n" + "="*80)
        print("FINAL CLEARANCE TEAM MONTHLY SCORECARD GENERATOR")
        print("="*80)
        
        self.load_production()
        self.load_attendance()
        self.load_pkt()
        self.load_internal_audit()
        self.load_final_clearance()
        self.load_client_audit()
        
        self.calculate_overall_score()
        scorecard = self.generate_scorecard()
        
        print("\n" + "="*80)
        print(f"✅ SCORECARD GENERATED FOR {len(scorecard)} FINAL CLEARANCE TEAM MEMBERS")
        print("="*80)
        
        return scorecard, self.employees

# Main execution
if __name__ == "__main__":
    folder_path = input("\n📁 Enter folder path containing Excel files: ").strip().strip('"')
    
    generator = FinalClearanceScorecard(folder_path)
    scorecard, employees_data = generator.run()
    
    # Save as JSON
    with open('/mnt/user-data/outputs/final_clearance_scorecard.json', 'w') as f:
        json.dump(scorecard, f, indent=2, default=str)
    
    print(f"\n✅ Scorecard saved to: final_clearance_scorecard.json")
    print(f"\n📊 Final Clearance Team Summary:")
    
    if scorecard:
        print(f"\n{'Employee Name':<25} {'Score':<8} {'Rating':<15} {'SLA %':<10} {'Incentive':<15}")
        print("-" * 80)
        for row in scorecard:
            print(f"{row['Employee Name']:<25} {str(row['Overall Score']):<8} {str(row['Overall Rating']):<15} {str(row['SLA Compliance %']):<10} {str(row['Incentive Eligible']):<15}")
