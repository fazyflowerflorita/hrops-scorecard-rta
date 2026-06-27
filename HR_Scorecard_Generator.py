#!/usr/bin/env python3
"""
HR Operations Monthly Scorecard Generator
Consolidated Analysis of 8 Excel Files
"""

import openpyxl
from datetime import datetime
import json
from pathlib import Path
from collections import defaultdict

class HRScorecardGenerator:
    def __init__(self, folder_path):
        self.folder = Path(folder_path)
        self.employees = defaultdict(dict)
        self.teams_map = {}
        
    def load_attendance(self):
        """Load Attendance Data"""
        print("\n📊 Loading Attendance...")
        file_path = self.folder / "Attendance.xlsx"
        
        if not file_path.exists():
            print(f"   ❌ File not found: {file_path}")
            return
        
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        
        current_month = datetime.now().month
        current_year = datetime.now().year
        
        for row_idx in range(2, ws.max_row + 1):
            emp_name = ws.cell(row=row_idx, column=1).value
            if not emp_name:
                continue
            
            emp_name = str(emp_name).strip()
            
            # Get current month leaves (column index = current_month + 1)
            leaves = ws.cell(row=row_idx, column=current_month + 1).value
            
            try:
                leaves = int(leaves) if leaves else 0
            except:
                leaves = 0
            
            # Calculate attendance %
            working_days = 22  # Approximate
            attendance_pct = max(0, ((working_days - leaves) / working_days) * 100)
            
            self.employees[emp_name]['attendance_leaves'] = leaves
            self.employees[emp_name]['attendance_pct'] = round(attendance_pct, 2)
            self.employees[emp_name]['attendance_status'] = 'Green' if leaves == 0 else 'Red' if leaves > 2 else 'Amber'
        
        wb.close()
        print(f"   ✅ Loaded {len([e for e in self.employees if 'attendance_leaves' in self.employees[e]])} employees")

    def load_production(self):
        """Load Production Tracker"""
        print("\n⚡ Loading Production...")
        file_path = self.folder / "Production_Tracker.xlsx"
        
        if not file_path.exists():
            print(f"   ❌ File not found: {file_path}")
            return
        
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        
        current_month = datetime.now().month
        
        for row_idx in range(2, ws.max_row + 1):
            emp_name = ws.cell(row=row_idx, column=1).value
            if not emp_name:
                continue
            
            emp_name = str(emp_name).strip()
            
            # Get current month production
            hours = ws.cell(row=row_idx, column=current_month + 1).value
            
            try:
                hours = float(hours) if hours else 0
            except:
                hours = 0
            
            productivity_status = 'Green' if hours >= 8 else 'Red'
            
            self.employees[emp_name]['productivity_hours'] = hours
            self.employees[emp_name]['productivity_target'] = 8
            self.employees[emp_name]['productivity_status'] = productivity_status
        
        wb.close()
        print(f"   ✅ Loaded production data")

    def load_pkt(self):
        """Load Process Knowledge Test"""
        print("\n📚 Loading Process Knowledge Test...")
        file_path = self.folder / "Process_Knowledge_Test.xlsx"
        
        if not file_path.exists():
            print(f"   ❌ File not found: {file_path}")
            return
        
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        
        current_month = datetime.now().month
        
        for row_idx in range(2, ws.max_row + 1):
            emp_name = ws.cell(row=row_idx, column=1).value
            if not emp_name:
                continue
            
            emp_name = str(emp_name).strip()
            
            # Get current month score
            score = ws.cell(row=row_idx, column=current_month + 1).value
            
            try:
                score = int(score) if score else 0
            except:
                score = 0
            
            if score >= 90:
                pkt_rating = 'Excellent'
                pkt_status = 'Green'
            elif score >= 80:
                pkt_rating = 'Good'
                pkt_status = 'Amber'
            else:
                pkt_rating = 'Needs Improvement'
                pkt_status = 'Red'
            
            self.employees[emp_name]['pkt_score'] = score
            self.employees[emp_name]['pkt_rating'] = pkt_rating
            self.employees[emp_name]['pkt_status'] = pkt_status
        
        wb.close()
        print(f"   ✅ Loaded PKT data")

    def load_nh_pending(self):
        """Load NH Pending Tracker"""
        print("\n📋 Loading NH Pending...")
        file_path = self.folder / "New_NH_pending_Tracker.xlsx"
        
        if not file_path.exists():
            print(f"   ❌ File not found: {file_path}")
            return
        
        wb = openpyxl.load_workbook(file_path)
        ws = wb['2026 NH Pending']
        
        nh_pending_count = defaultdict(int)
        nh_candidates = defaultdict(list)
        
        for row_idx in range(2, ws.max_row + 1):
            status = ws.cell(row=row_idx, column=12).value  # Status column
            processor = ws.cell(row=row_idx, column=7).value  # Processor column
            candidate = ws.cell(row=row_idx, column=3).value  # Candidate Name
            
            if status and 'Pending' in str(status):
                if processor:
                    processor = str(processor).strip()
                    nh_pending_count[processor] += 1
                    if candidate:
                        nh_candidates[processor].append(str(candidate).strip())
        
        for emp_name in nh_pending_count:
            count = nh_pending_count[emp_name]
            self.employees[emp_name]['nh_pending_count'] = count
            self.employees[emp_name]['nh_eligibility'] = 'Not Eligible' if count > 0 else 'Eligible for Incentive'
            self.employees[emp_name]['nh_status'] = 'Red' if count > 0 else 'Green'
        
        wb.close()
        print(f"   ✅ Loaded NH Pending data")

    def load_client_audit(self):
        """Load Client System Audit"""
        print("\n🔍 Loading Client System Audit...")
        file_path = self.folder / "Client_System_Audit_Tracker.xlsx"
        
        if not file_path.exists():
            print(f"   ❌ File not found: {file_path}")
            return
        
        wb = openpyxl.load_workbook(file_path)
        ws = wb['Client System Audit']
        
        audit_pending = defaultdict(int)
        
        for row_idx in range(2, ws.max_row + 1):
            team_member = ws.cell(row=row_idx, column=3).value  # Team Member column
            pending = ws.cell(row=row_idx, column=8).value  # Pending column
            
            if team_member and pending:
                team_member = str(team_member).strip()
                try:
                    audit_pending[team_member] += int(pending)
                except:
                    pass
        
        for emp_name in audit_pending:
            count = audit_pending[emp_name]
            self.employees[emp_name]['client_audit_pending'] = count
            self.employees[emp_name]['client_audit_eligibility'] = 'Not Eligible' if count > 0 else 'Eligible for Incentive'
            self.employees[emp_name]['client_audit_status'] = 'Red' if count > 0 else 'Green'
        
        wb.close()
        print(f"   ✅ Loaded Client Audit data")

    def load_data_changes(self):
        """Load Data Changes Tracker"""
        print("\n📝 Loading Data Changes...")
        file_path = self.folder / "Data_Changes_Tracker.xlsx"
        
        if not file_path.exists():
            print(f"   ❌ File not found: {file_path}")
            return
        
        wb = openpyxl.load_workbook(file_path)
        ws = wb['Data change Completed']
        
        data_changes_status = defaultdict(lambda: {'total': 0, 'completed': 0})
        
        for row_idx in range(2, ws.max_row + 1):
            assigned_to = ws.cell(row=row_idx, column=7).value  # Assigned to column
            completed_date = ws.cell(row=row_idx, column=8).value  # Completed Date
            
            if assigned_to:
                assigned_to = str(assigned_to).strip()
                data_changes_status[assigned_to]['total'] += 1
                if completed_date:
                    data_changes_status[assigned_to]['completed'] += 1
        
        for emp_name in data_changes_status:
            total = data_changes_status[emp_name]['total']
            completed = data_changes_status[emp_name]['completed']
            completion_pct = (completed / total * 100) if total > 0 else 0
            
            self.employees[emp_name]['data_changes_total'] = total
            self.employees[emp_name]['data_changes_completed'] = completed
            self.employees[emp_name]['data_changes_status'] = 'Green' if completion_pct >= 90 else 'Red' if completion_pct < 70 else 'Amber'
        
        wb.close()
        print(f"   ✅ Loaded Data Changes data")

    def load_tenure_discount(self):
        """Load Tenure Discount Audit"""
        print("\n🎯 Loading Tenure Discount...")
        file_path = self.folder / "Tenure_Discount_Audit_Tracker.xlsx"
        
        if not file_path.exists():
            print(f"   ❌ File not found: {file_path}")
            return
        
        wb = openpyxl.load_workbook(file_path)
        ws = wb['Tenure discount audit tracker']
        
        for row_idx in range(2, ws.max_row + 1):
            team_member = ws.cell(row=row_idx, column=3).value  # Team Member column
            pending = ws.cell(row=row_idx, column=6).value  # Pending column
            
            if team_member:
                team_member = str(team_member).strip()
                try:
                    pending_count = int(pending) if pending else 0
                except:
                    pending_count = 0
                
                self.employees[team_member]['tenure_pending'] = pending_count
                self.employees[team_member]['tenure_status'] = 'Red' if pending_count > 0 else 'Green'
        
        wb.close()
        print(f"   ✅ Loaded Tenure Discount data")

    def load_termination(self):
        """Load Termination Tracker"""
        print("\n⚠️  Loading Termination...")
        file_path = self.folder / "Termination_Tracker.xlsx"
        
        if not file_path.exists():
            print(f"   ❌ File not found: {file_path}")
            return
        
        wb = openpyxl.load_workbook(file_path)
        
        for sheet_name in ['Termination RTA', 'Termination Pnow']:
            if sheet_name not in wb.sheetnames:
                continue
            
            ws = wb[sheet_name]
            
            for row_idx in range(2, ws.max_row + 1):
                end_date = ws.cell(row=row_idx, column=9).value  # End Date
                notified_date = ws.cell(row=row_idx, column=2).value  # Notified Date
                
                if end_date and notified_date:
                    try:
                        end_dt = end_date if isinstance(end_date, datetime) else datetime.strptime(str(end_date), '%Y-%m-%d')
                        notified_dt = notified_date if isinstance(notified_date, datetime) else datetime.strptime(str(notified_date), '%Y-%m-%d')
                        days_diff = (notified_dt - end_dt).days
                        
                        # Store the most recent termination
                        if 'termination_days' not in self.employees or days_diff > self.employees.get('termination_days', 0):
                            pass  # Can be enhanced with proper employee tracking
                    except:
                        pass
        
        wb.close()
        print(f"   ✅ Loaded Termination data")

    def calculate_overall_score(self):
        """Calculate Overall Score for each employee"""
        print("\n🎯 Calculating Overall Scores...")
        
        for emp_name in self.employees:
            emp = self.employees[emp_name]
            
            scores = []
            weights = []
            
            # Quality (40%)
            if 'pkt_score' in emp:
                scores.append(emp['pkt_score'])
                weights.append(40)
            
            # Attendance (20%)
            if 'attendance_pct' in emp:
                scores.append(emp['attendance_pct'])
                weights.append(20)
            
            # Productivity (20%)
            if 'productivity_hours' in emp:
                prod_score = min((emp['productivity_hours'] / 8) * 100, 100)
                scores.append(prod_score)
                weights.append(20)
            
            # Calculate weighted average
            if scores:
                total_weight = sum(weights)
                overall = sum(s * w for s, w in zip(scores, weights)) / total_weight if total_weight > 0 else 0
                emp['overall_score'] = round(overall, 2)
                
                # Determine overall rating
                if overall >= 90:
                    emp['overall_rating'] = 'Excellent'
                    emp['overall_status'] = 'Green'
                elif overall >= 80:
                    emp['overall_rating'] = 'Good'
                    emp['overall_status'] = 'Amber'
                else:
                    emp['overall_rating'] = 'Needs Improvement'
                    emp['overall_status'] = 'Red'
                
                # Check incentive eligibility
                nh_pending = emp.get('nh_pending_count', 0)
                audit_pending = emp.get('client_audit_pending', 0)
                
                if nh_pending > 0 or audit_pending > 0:
                    emp['incentive_eligible'] = 'Not Eligible'
                else:
                    emp['incentive_eligible'] = 'Eligible'

    def generate_report(self):
        """Generate consolidated scorecard"""
        print("\n📊 Generating Report...")
        
        report = []
        for emp_name in sorted(self.employees.keys()):
            emp = self.employees[emp_name]
            
            report_row = {
                'Employee Name': emp_name,
                'Productivity Hours': emp.get('productivity_hours', '-'),
                'Productivity Status': emp.get('productivity_status', '-'),
                'Attendance %': emp.get('attendance_pct', '-'),
                'Attendance Status': emp.get('attendance_status', '-'),
                'PKT Score': emp.get('pkt_score', '-'),
                'PKT Rating': emp.get('pkt_rating', '-'),
                'NH Pending': emp.get('nh_pending_count', 0),
                'NH Eligibility': emp.get('nh_eligibility', '-'),
                'Audit Pending': emp.get('client_audit_pending', 0),
                'Audit Eligibility': emp.get('client_audit_eligibility', '-'),
                'Data Changes': emp.get('data_changes_completed', '-'),
                'Tenure Pending': emp.get('tenure_pending', 0),
                'Overall Score': emp.get('overall_score', '-'),
                'Overall Rating': emp.get('overall_rating', '-'),
                'Incentive Eligible': emp.get('incentive_eligible', '-'),
            }
            report.append(report_row)
        
        return report

    def run(self):
        """Execute all data loading"""
        print("\n" + "="*80)
        print("HR OPERATIONS SCORECARD GENERATOR")
        print("="*80)
        
        self.load_attendance()
        self.load_production()
        self.load_pkt()
        self.load_nh_pending()
        self.load_client_audit()
        self.load_data_changes()
        self.load_tenure_discount()
        self.load_termination()
        
        self.calculate_overall_score()
        report = self.generate_report()
        
        print("\n" + "="*80)
        print(f"SCORECARD GENERATED FOR {len(report)} EMPLOYEES")
        print("="*80)
        
        return report, self.employees

# Main execution
if __name__ == "__main__":
    folder_path = input("📁 Enter folder path containing Excel files: ").strip().strip('"')
    
    generator = HRScorecardGenerator(folder_path)
    report, employees_data = generator.run()
    
    # Save report as JSON
    with open('hr_scorecard_report.json', 'w') as f:
        json.dump(report, f, indent=2, default=str)
    
    print(f"\n✅ Report saved to: hr_scorecard_report.json")
    print(f"\n📊 Sample Report:")
    for row in report[:3]:
        print(f"\n   {row['Employee Name']}")
        print(f"      Score: {row['Overall Score']} ({row['Overall Rating']})")
        print(f"      PKT: {row['PKT Score']} | Attendance: {row['Attendance %']}%")
