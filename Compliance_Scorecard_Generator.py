#!/usr/bin/env python3
"""
Compliance Team Monthly Scorecard Generator
Analyzes 6 Excel files and generates consolidated performance scorecard
"""

import openpyxl
from datetime import datetime
import json
from pathlib import Path
from collections import defaultdict
import re

class ComplianceScorecard:
    def __init__(self, folder_path):
        self.folder = Path(folder_path)
        self.employees = defaultdict(dict)
        self.current_month = datetime.now().month
        self.compliance_team = [
            'Sayee Nivas B', 'Alan Benjamin', 'Pavithra M', 'Latha J', 
            'Sneha Thomas', 'Azhar Taj', 'Rathina Sudhan K'
        ]
        
    def normalize_name(self, name):
        """Normalize employee names for matching"""
        if not name:
            return ""
        name = str(name).strip().lower()
        # Remove extra spaces and standardize
        name = ' '.join(name.split())
        return name
    
    def is_compliance_team(self, name):
        """Check if employee is in Compliance Team"""
        if not name:
            return False
        name_norm = self.normalize_name(name)
        for team_member in self.compliance_team:
            if self.normalize_name(team_member) in name_norm or name_norm in self.normalize_name(team_member):
                return True
        return False
    
    def load_production(self):
        """Load Production Tracker Data"""
        print("\n⚡ Loading Production Tracker...")
        file_path = self.folder / "Production_Tracker.xlsx"
        
        if not file_path.exists():
            print(f"   ❌ File not found")
            return
        
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        
        for row_idx in range(2, ws.max_row + 1):
            emp_name = ws.cell(row=row_idx, column=1).value
            if not emp_name or not self.is_compliance_team(emp_name):
                continue
            
            emp_name = str(emp_name).strip()
            hours = ws.cell(row=row_idx, column=self.current_month + 1).value
            
            try:
                hours = float(hours) if hours else 0
            except:
                hours = 0
            
            status = 'Green' if hours >= 8 else 'Red'
            
            self.employees[emp_name]['productivity_hours'] = hours
            self.employees[emp_name]['productivity_target'] = 8
            self.employees[emp_name]['productivity_status'] = status
        
        wb.close()
        print(f"   ✅ Loaded production data")
    
    def load_attendance(self):
        """Load Attendance Data"""
        print("\n📊 Loading Attendance...")
        file_path = self.folder / "Attendance.xlsx"
        
        if not file_path.exists():
            print(f"   ❌ File not found")
            return
        
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        
        for row_idx in range(2, ws.max_row + 1):
            emp_name = ws.cell(row=row_idx, column=1).value
            if not emp_name or not self.is_compliance_team(emp_name):
                continue
            
            emp_name = str(emp_name).strip()
            leaves = ws.cell(row=row_idx, column=self.current_month + 1).value
            
            try:
                leaves = int(leaves) if leaves else 0
            except:
                leaves = 0
            
            working_days = 22
            attendance_pct = max(0, ((working_days - leaves) / working_days) * 100)
            
            self.employees[emp_name]['attendance_leaves'] = leaves
            self.employees[emp_name]['attendance_pct'] = round(attendance_pct, 2)
            self.employees[emp_name]['attendance_status'] = 'Green' if leaves == 0 else 'Red' if leaves > 2 else 'Amber'
        
        wb.close()
        print(f"   ✅ Loaded attendance data")
    
    def load_pkt(self):
        """Load Process Knowledge Test"""
        print("\n📚 Loading Process Knowledge Test...")
        file_path = self.folder / "Process_Knowledge_Test.xlsx"
        
        if not file_path.exists():
            print(f"   ❌ File not found")
            return
        
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        
        for row_idx in range(2, ws.max_row + 1):
            emp_name = ws.cell(row=row_idx, column=1).value
            if not emp_name or not self.is_compliance_team(emp_name):
                continue
            
            emp_name = str(emp_name).strip()
            score = ws.cell(row=row_idx, column=self.current_month + 1).value
            
            try:
                score = int(score) if score else 0
            except:
                score = 0
            
            if score >= 90:
                rating = 'Excellent'
                status = 'Green'
            elif score >= 80:
                rating = 'Good'
                status = 'Amber'
            else:
                rating = 'Needs Improvement'
                status = 'Red'
            
            self.employees[emp_name]['pkt_score'] = score
            self.employees[emp_name]['pkt_rating'] = rating
            self.employees[emp_name]['pkt_status'] = status
        
        wb.close()
        print(f"   ✅ Loaded PKT data")
    
    def load_internal_audit(self):
        """Load Internal Audit Errors"""
        print("\n🔍 Loading Internal Audit...")
        file_path = self.folder / "Internal_Audit_Scores.xlsx"
        
        if not file_path.exists():
            print(f"   ❌ File not found")
            return
        
        wb = openpyxl.load_workbook(file_path)
        
        # Parse Internal Audit sheet for monthly errors
        if 'Internal Audit' in wb.sheetnames:
            ws = wb['Internal Audit']
            
            error_count = defaultdict(int)
            
            for row_idx in range(2, ws.max_row + 1):
                team_member_cell = ws.cell(row=row_idx, column=5).value
                error_cell = ws.cell(row=row_idx, column=2).value
                
                if team_member_cell and error_cell:
                    # Parse error count and team members
                    try:
                        errors = int(str(error_cell).split()[0])
                        # Parse team members from string like "Sayee (1) & Pavithra"
                        members_str = str(team_member_cell)
                        for comp_member in self.compliance_team:
                            if self.normalize_name(comp_member) in members_str.lower():
                                error_count[comp_member] += errors
                    except:
                        pass
        
        # Parse Error Tracker sheet for auditor/processor errors
        if 'Error Tracker' in wb.sheetnames:
            ws = wb['Error Tracker']
            
            for row_idx in range(2, ws.max_row + 1):
                auditor = ws.cell(row=row_idx, column=9).value
                processor = ws.cell(row=row_idx, column=7).value
                error_count_cell = ws.cell(row=row_idx, column=11).value
                
                if error_count_cell:
                    try:
                        errors = int(error_count_cell) if error_count_cell else 0
                        if auditor and self.is_compliance_team(auditor):
                            auditor = str(auditor).strip()
                            error_count[auditor] += errors
                        if processor and self.is_compliance_team(processor):
                            processor = str(processor).strip()
                            error_count[processor] += errors
                    except:
                        pass
        
        for emp_name in error_count:
            errors = error_count[emp_name]
            if errors == 0:
                status = 'Green'
            elif errors <= 2:
                status = 'Amber'
            else:
                status = 'Red'
            
            self.employees[emp_name]['audit_errors'] = errors
            self.employees[emp_name]['audit_status'] = status
        
        wb.close()
        print(f"   ✅ Loaded internal audit data")
    
    def load_nh_pending(self):
        """Load NH Pending Data"""
        print("\n📋 Loading NH Pending...")
        file_path = self.folder / "New_NH_pending_Tracker.xlsx"
        
        if not file_path.exists():
            print(f"   ❌ File not found")
            return
        
        wb = openpyxl.load_workbook(file_path)
        ws = wb['2026 NH Pending']
        
        nh_pending = defaultdict(int)
        
        for row_idx in range(2, ws.max_row + 1):
            status = ws.cell(row=row_idx, column=12).value
            processor = ws.cell(row=row_idx, column=7).value
            
            if status and 'Pending' in str(status):
                if processor and self.is_compliance_team(processor):
                    processor = str(processor).strip()
                    nh_pending[processor] += 1
        
        for emp_name in nh_pending:
            count = nh_pending[emp_name]
            self.employees[emp_name]['nh_pending_count'] = count
            self.employees[emp_name]['nh_eligibility'] = 'Not Eligible' if count > 0 else 'Eligible'
            self.employees[emp_name]['nh_status'] = 'Red' if count > 0 else 'Green'
        
        wb.close()
        print(f"   ✅ Loaded NH pending data")
    
    def load_client_audit(self):
        """Load Client System Audit"""
        print("\n🏢 Loading Client System Audit...")
        file_path = self.folder / "Client_System_Audit_Tracker.xlsx"
        
        if not file_path.exists():
            print(f"   ❌ File not found")
            return
        
        wb = openpyxl.load_workbook(file_path)
        ws = wb['Client System Audit']
        
        audit_pending = defaultdict(int)
        
        for row_idx in range(2, ws.max_row + 1):
            team_member = ws.cell(row=row_idx, column=3).value
            pending = ws.cell(row=row_idx, column=8).value
            
            if team_member and self.is_compliance_team(team_member):
                team_member = str(team_member).strip()
                try:
                    pending_count = int(pending) if pending else 0
                    audit_pending[team_member] += pending_count
                except:
                    pass
        
        for emp_name in audit_pending:
            count = audit_pending[emp_name]
            self.employees[emp_name]['client_audit_pending'] = count
            self.employees[emp_name]['client_audit_eligibility'] = 'Not Eligible' if count > 0 else 'Eligible'
            self.employees[emp_name]['client_audit_status'] = 'Red' if count > 0 else 'Green'
        
        wb.close()
        print(f"   ✅ Loaded client audit data")
    
    def calculate_overall_score(self):
        """Calculate Overall Score and Rating"""
        print("\n🎯 Calculating Overall Scores...")
        
        for emp_name in self.employees:
            emp = self.employees[emp_name]
            
            scores = []
            weights = []
            
            # PKT: 40%
            if 'pkt_score' in emp:
                scores.append(emp['pkt_score'])
                weights.append(40)
            
            # Attendance: 20%
            if 'attendance_pct' in emp:
                scores.append(emp['attendance_pct'])
                weights.append(20)
            
            # Productivity: 20%
            if 'productivity_hours' in emp:
                prod_score = min((emp['productivity_hours'] / 8) * 100, 100)
                scores.append(prod_score)
                weights.append(20)
            
            # Internal Audit: 20% (inverse - fewer errors = higher score)
            if 'audit_errors' in emp:
                audit_score = max(0, 100 - (emp['audit_errors'] * 10))
                scores.append(audit_score)
                weights.append(20)
            
            if scores:
                total_weight = sum(weights)
                overall = sum(s * w for s, w in zip(scores, weights)) / total_weight if total_weight > 0 else 0
                emp['overall_score'] = round(overall, 2)
                
                if overall >= 90:
                    emp['overall_rating'] = 'Excellent'
                    emp['overall_status'] = 'Green'
                elif overall >= 80:
                    emp['overall_rating'] = 'Good'
                    emp['overall_status'] = 'Amber'
                else:
                    emp['overall_rating'] = 'Needs Improvement'
                    emp['overall_status'] = 'Red'
                
                # Incentive eligibility
                nh_pending = emp.get('nh_pending_count', 0)
                audit_pending = emp.get('client_audit_pending', 0)
                
                if nh_pending > 0 or audit_pending > 0:
                    emp['incentive_eligible'] = 'Not Eligible'
                else:
                    emp['incentive_eligible'] = 'Eligible'
    
    def generate_scorecard(self):
        """Generate consolidated scorecard"""
        print("\n📊 Generating Compliance Team Scorecard...")
        
        scorecard = []
        for emp_name in sorted(self.employees.keys()):
            emp = self.employees[emp_name]
            
            record = {
                'Employee Name': emp_name,
                'Productivity Hours': emp.get('productivity_hours', '-'),
                'Productivity Status': emp.get('productivity_status', '-'),
                'PKT Score': emp.get('pkt_score', '-'),
                'PKT Rating': emp.get('pkt_rating', '-'),
                'Attendance Leaves': emp.get('attendance_leaves', '-'),
                'Attendance %': emp.get('attendance_pct', '-'),
                'Attendance Status': emp.get('attendance_status', '-'),
                'Audit Errors': emp.get('audit_errors', 0),
                'Audit Status': emp.get('audit_status', '-'),
                'NH Pending': emp.get('nh_pending_count', 0),
                'NH Eligibility': emp.get('nh_eligibility', '-'),
                'Client Audit Pending': emp.get('client_audit_pending', 0),
                'Client Audit Eligibility': emp.get('client_audit_eligibility', '-'),
                'Overall Score': emp.get('overall_score', '-'),
                'Overall Rating': emp.get('overall_rating', '-'),
                'Incentive Eligible': emp.get('incentive_eligible', '-'),
                'Remarks': 'On Track' if emp.get('incentive_eligible') == 'Eligible' else 'Action Required'
            }
            scorecard.append(record)
        
        return scorecard
    
    def run(self):
        """Execute all data loading and processing"""
        print("\n" + "="*80)
        print("COMPLIANCE TEAM MONTHLY SCORECARD GENERATOR")
        print("="*80)
        
        self.load_production()
        self.load_attendance()
        self.load_pkt()
        self.load_internal_audit()
        self.load_nh_pending()
        self.load_client_audit()
        
        self.calculate_overall_score()
        scorecard = self.generate_scorecard()
        
        print("\n" + "="*80)
        print(f"✅ SCORECARD GENERATED FOR {len(scorecard)} COMPLIANCE TEAM MEMBERS")
        print("="*80)
        
        return scorecard, self.employees

# Main execution
if __name__ == "__main__":
    folder_path = input("\n📁 Enter folder path containing Excel files: ").strip().strip('"')
    
    generator = ComplianceScorecard(folder_path)
    scorecard, employees_data = generator.run()
    
    # Save as JSON
    with open('/mnt/user-data/outputs/compliance_scorecard.json', 'w') as f:
        json.dump(scorecard, f, indent=2, default=str)
    
    print(f"\n✅ Scorecard saved to: compliance_scorecard.json")
    print(f"\n📊 Compliance Team Summary:")
    
    if scorecard:
        print(f"\n{'Employee Name':<25} {'Score':<8} {'Rating':<15} {'Incentive':<15}")
        print("-" * 70)
        for row in scorecard:
            print(f"{row['Employee Name']:<25} {str(row['Overall Score']):<8} {str(row['Overall Rating']):<15} {str(row['Incentive Eligible']):<15}")
