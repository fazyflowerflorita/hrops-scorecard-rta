#!/usr/bin/env python3
"""
HR Operations Scorecard - Enhanced Multi-File Parser
Handles all 37+ Excel files intelligently
Auto-syncs to Firebase
"""

import os
import sys
import time
from pathlib import Path
import openpyxl
import firebase_admin
from firebase_admin import credentials, db
from datetime import datetime
import hashlib
import json

FIREBASE_CONFIG = {
    "projectId": "hrops-scorecard---rta",
    "databaseURL": "https://hrops-scorecard---rta-default-rtdb.firebaseio.com"
}

class EnhancedExcelParser:
    """Parse all Excel files intelligently"""
    
    @staticmethod
    def read_excel(file_path):
        """Read Excel file safely"""
        try:
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            
            headers = []
            for cell in ws[1]:
                if cell.value:
                    headers.append(str(cell.value).strip())
            
            data = []
            for row in ws.iter_rows(min_row=2, values_only=False):
                row_data = {}
                for idx, cell in enumerate(row):
                    if idx < len(headers):
                        row_data[headers[idx]] = cell.value
                if any(row_data.values()):
                    data.append(row_data)
            
            wb.close()
            return data, headers
        except Exception as e:
            return [], []
    
    @staticmethod
    def extract_employee_id(row):
        """Extract employee ID from any row"""
        for key, value in row.items():
            key_lower = str(key).lower()
            if 'emp' in key_lower and 'id' in key_lower and value:
                return str(value).strip().upper()
            if 'id' in key_lower and value and len(str(value)) < 20:
                val = str(value).strip().upper()
                if val.startswith('P') or val.isdigit():
                    return val
        return None
    
    @staticmethod
    def extract_name(row):
        """Extract employee name from any row"""
        for key, value in row.items():
            key_lower = str(key).lower()
            if 'name' in key_lower and value:
                return str(value).strip()
            if 'employee' in key_lower and 'name' not in key_lower and value:
                return str(value).strip()
        return None
    
    @staticmethod
    def extract_number(value):
        """Safe number extraction"""
        if not value:
            return 0
        try:
            val_str = str(value).strip().replace('%', '')
            return float(val_str)
        except:
            return 0
    
    @staticmethod
    def parse_file(file_path):
        """Intelligently parse file based on name and content"""
        file_name = Path(file_path).name.lower()
        data, headers = EnhancedExcelParser.read_excel(file_path)
        
        if not data:
            return {}
        
        result = {}
        
        # ATTENDANCE FILES
        if 'attendance' in file_name:
            for row in data:
                emp_id = EnhancedExcelParser.extract_employee_id(row)
                if not emp_id:
                    continue
                
                attendance = 0
                for key, value in row.items():
                    key_lower = str(key).lower()
                    if 'attend' in key_lower or '%' in str(value or ''):
                        attendance = EnhancedExcelParser.extract_number(value)
                        break
                
                if emp_id not in result:
                    result[emp_id] = {}
                result[emp_id]['attendance'] = attendance
        
        # PROCESS KNOWLEDGE FILES
        elif 'knowledge' in file_name or 'process knowledge' in file_name:
            for row in data:
                emp_id = EnhancedExcelParser.extract_employee_id(row)
                if not emp_id:
                    continue
                
                score = 0
                for key, value in row.items():
                    key_lower = str(key).lower()
                    if 'score' in key_lower or 'test' in key_lower:
                        score = EnhancedExcelParser.extract_number(value)
                        break
                
                if emp_id not in result:
                    result[emp_id] = {}
                result[emp_id]['process_knowledge'] = score
        
        # PRODUCTION TRACKER
        elif 'production' in file_name:
            for row in data:
                emp_id = EnhancedExcelParser.extract_employee_id(row)
                if not emp_id:
                    continue
                
                productivity = 0
                for key, value in row.items():
                    key_lower = str(key).lower()
                    if 'count' in key_lower or 'completed' in key_lower or 'produced' in key_lower:
                        productivity = EnhancedExcelParser.extract_number(value)
                        break
                
                if emp_id not in result:
                    result[emp_id] = {}
                result[emp_id]['productivity'] = productivity
        
        # AUDIT FILES (Client System, Internal Audit, etc.)
        elif any(x in file_name for x in ['audit', 'qmg error', 'error tracker']):
            for row in data:
                emp_id = EnhancedExcelParser.extract_employee_id(row)
                if not emp_id:
                    continue
                
                score = 0
                count = 0
                for key, value in row.items():
                    key_lower = str(key).lower()
                    if 'score' in key_lower or 'rating' in key_lower:
                        score = EnhancedExcelParser.extract_number(value)
                    if 'error' in key_lower or 'finding' in key_lower or 'count' in key_lower:
                        count = EnhancedExcelParser.extract_number(value)
                
                if emp_id not in result:
                    result[emp_id] = {}
                result[emp_id]['audit_score'] = score
                result[emp_id]['audit_errors'] = count
        
        # NH PENDING TRACKER
        elif 'nh pending' in file_name or 'bg pending' in file_name:
            for row in data:
                emp_id = EnhancedExcelParser.extract_employee_id(row)
                if not emp_id:
                    continue
                
                pending = 0
                for key, value in row.items():
                    key_lower = str(key).lower()
                    if 'pending' in key_lower or 'count' in key_lower:
                        pending = EnhancedExcelParser.extract_number(value)
                        break
                
                if emp_id not in result:
                    result[emp_id] = {}
                result[emp_id]['nh_pending'] = pending
        
        # PAPERWORK TRACKER
        elif 'paperwork' in file_name:
            for row in data:
                emp_id = EnhancedExcelParser.extract_employee_id(row)
                if not emp_id:
                    continue
                
                paperwork_count = 0
                for key, value in row.items():
                    key_lower = str(key).lower()
                    if 'count' in key_lower or 'pending' in key_lower or 'allocated' in key_lower:
                        paperwork_count = EnhancedExcelParser.extract_number(value)
                        break
                
                if emp_id not in result:
                    result[emp_id] = {}
                result[emp_id]['paperwork'] = paperwork_count
        
        # TERMINATION TRACKER
        elif 'termination' in file_name:
            for row in data:
                emp_id = EnhancedExcelParser.extract_employee_id(row)
                if not emp_id:
                    continue
                
                if emp_id not in result:
                    result[emp_id] = {}
                result[emp_id]['status'] = 'Terminated'
        
        # DATA CHANGES TRACKER
        elif 'data changes' in file_name or 'changes tracker' in file_name:
            for row in data:
                emp_id = EnhancedExcelParser.extract_employee_id(row)
                if not emp_id:
                    continue
                
                changes = 0
                for key, value in row.items():
                    key_lower = str(key).lower()
                    if 'change' in key_lower or 'count' in key_lower:
                        changes = EnhancedExcelParser.extract_number(value)
                        break
                
                if emp_id not in result:
                    result[emp_id] = {}
                result[emp_id]['data_changes'] = changes
        
        return result

class DataConsolidator:
    """Consolidate data from all files"""
    
    @staticmethod
    def consolidate(all_file_data):
        """Merge all file data into unified employee records"""
        consolidated = {}
        
        for file_data in all_file_data:
            for emp_id, metrics in file_data.items():
                if emp_id not in consolidated:
                    consolidated[emp_id] = {
                        'empId': emp_id,
                        'lastUpdated': datetime.now().isoformat()
                    }
                consolidated[emp_id].update(metrics)
        
        return consolidated

class FirebaseSync:
    """Sync data to Firebase"""
    
    @staticmethod
    def initialize():
        """Initialize Firebase"""
        try:
            if firebase_admin._apps:
                firebase_admin.delete_app(firebase_admin.get_app())
            
            cred = credentials.Certificate('firebase-service-account.json')
            firebase_admin.initialize_app(cred, {
                'databaseURL': FIREBASE_CONFIG['databaseURL']
            })
            return True
        except Exception as e:
            print(f"❌ Firebase error: {e}")
            return False
    
    @staticmethod
    def sync_data(consolidated_data):
        """Sync to Firebase"""
        try:
            now = datetime.now().isoformat()
            employee_master = {}
            performance_data = {}

            for emp_id, record in consolidated_data.items():
                name = record.get('name') or emp_id.replace('_', ' ').title()
                attendance = EnhancedExcelParser.extract_number(record.get('attendance'))
                knowledge = EnhancedExcelParser.extract_number(record.get('process_knowledge'))
                productivity = EnhancedExcelParser.extract_number(record.get('productivity'))
                audit_score = EnhancedExcelParser.extract_number(record.get('audit_score'))
                score_inputs = [value for value in [attendance, knowledge, audit_score] if value > 0]
                overall_score = round(sum(score_inputs) / len(score_inputs), 2) if score_inputs else 0

                employee_master[emp_id] = {
                    'empId': emp_id,
                    'name': name,
                    'team': record.get('team', 'Unassigned'),
                    'manager': record.get('manager', ''),
                    'department': record.get('department', 'HR Operations'),
                    'lastUpdated': now
                }

                performance_data[emp_id] = {
                    **record,
                    'empId': emp_id,
                    'name': name,
                    'attendance': attendance,
                    'process_knowledge': knowledge,
                    'productivity': productivity,
                    'audit_score': audit_score,
                    'overall_score': overall_score,
                    'lastUpdated': now
                }

            db.reference('/').update({
                'employee-master': employee_master,
                'performance-data': performance_data,
                'employee-module-data': performance_data,
                'employee-scores': performance_data,
                'sync-status': {
                    'lastUpdated': now,
                    'employeeCount': len(performance_data)
                }
            })
            return True
        except Exception as e:
            print(f"❌ Sync error: {e}")
            return False

class FolderMonitor:
    """Monitor folder for Excel files"""
    
    def __init__(self, folder_path):
        self.folder_path = Path(folder_path)
        self.file_hashes = {}
        self.last_sync = None
    
    def get_file_hash(self, file_path):
        """Get file hash"""
        try:
            with open(file_path, 'rb') as f:
                return hashlib.md5(f.read()).hexdigest()
        except:
            return None
    
    def find_excel_files(self):
        """Find all Excel files"""
        files = []
        for f in self.folder_path.glob('*.xlsx'):
            if not f.name.startswith('~$'):
                files.append(f)
        return files
    
    def process_files(self):
        """Process all Excel files"""
        excel_files = self.find_excel_files()
        
        if not excel_files:
            return False
        
        print(f"\n📁 Found {len(excel_files)} Excel files")
        
        all_data = []
        file_count = 0
        any_changed = False
        
        for file_path in excel_files:
            current_hash = self.get_file_hash(file_path)
            file_key = str(file_path)

            # Check if any file changed. When one file changes, re-parse every workbook
            # so Firebase receives a complete current scorecard snapshot.
            if file_key not in self.file_hashes or self.file_hashes[file_key] != current_hash:
                self.file_hashes[file_key] = current_hash
                any_changed = True

        if not any_changed:
            return False

        for file_path in excel_files:
            print(f"  Processing: {file_path.name}")
            file_data = EnhancedExcelParser.parse_file(str(file_path))

            if file_data:
                all_data.append(file_data)
                file_count += 1

        if not all_data:
            return False
        
        # Consolidate
        print(f"\n🔗 Consolidating data from {file_count} files...")
        consolidated = DataConsolidator.consolidate(all_data)
        
        # Sync to Firebase
        print(f"☁️ Syncing {len(consolidated)} employees to Firebase...")
        if FirebaseSync.sync_data(consolidated):
            self.last_sync = datetime.now()
            print(f"✅ Sync complete! Last update: {self.last_sync.strftime('%Y-%m-%d %H:%M:%S')}")
            return True
        
        return False
    
    def start(self, interval=30):
        """Start monitoring"""
        print("\n╔════════════════════════════════════════════════════════════════╗")
        print("║   HR OPERATIONS SCORECARD - Multi-File Monitor                ║")
        print("║   Processing all Excel files automatically                    ║")
        print("╚════════════════════════════════════════════════════════════════╝\n")
        
        print(f"📁 Monitoring: {self.folder_path}")
        print(f"⏱️  Checking every {interval} seconds")
        print("Press Ctrl+C to stop...\n")
        
        try:
            iteration = 0
            while True:
                iteration += 1
                print(f"\n[{datetime.now().strftime('%H:%M:%S')}] Check #{iteration}")
                self.process_files()
                time.sleep(interval)
        except KeyboardInterrupt:
            print("\n\n✅ Monitor stopped.")

def main():
    print("\nEnter your local folder path:")
    print("Example: C:\\Users\\FazyFlowerFlorita\\Pride Technologies\\RTA - 2026\n")
    
    folder_path = sys.argv[1].strip().strip('"') if len(sys.argv) > 1 else input("Folder path: ").strip().strip('"')
    
    if not folder_path or not Path(folder_path).exists():
        print("❌ Folder not found!")
        return
    
    # Initialize Firebase
    print("\n🔄 Initializing Firebase...")
    if not FirebaseSync.initialize():
        print("❌ Firebase initialization failed!")
        return
    
    print("✅ Firebase connected!")
    
    # Start monitoring
    monitor = FolderMonitor(folder_path)
    monitor.start(interval=30)

if __name__ == "__main__":
    main()
