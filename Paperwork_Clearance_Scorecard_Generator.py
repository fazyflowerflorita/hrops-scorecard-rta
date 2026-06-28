#!/usr/bin/env python3
"""
Paperwork Clearance Team Monthly Scorecard Generator
Analyzes 8 Excel files including Paperwork SLA & Allocation tracking
"""

import openpyxl
from datetime import datetime, timedelta
import json
from pathlib import Path
from collections import defaultdict

class PaperworkClearanceScorecard:
    def __init__(self, folder_path):
        self.folder = Path(folder_path)
        self.employees = defaultdict(dict)
        self.current_month = datetime.now().month
        self.paperwork_team = [
            'Leonie Gomes',
            'Thirisha Manoharan',
            'Vinish Navinkumar'
        ]
        
    def normalize_name(self, name):
        if not name:
            return ""
        return str(name).strip().lower()
    
    def is_paperwork_team(self, name):
        if not name:
            return False
        name_norm = self.normalize_name(name)
        for member in self.paperwork_team:
            if self.normalize_name(member) in name_norm or name_norm in self.normalize_name(member):
                return True
        return False
    
    def load_production(self):
        print("\n⚡ Loading Production...")
        file_path = self.folder / "Production_Tracker.xlsx"
        if not file_path.exists():
            return
        
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        
        for row_idx in range(2, ws.max_row + 1):
            emp_name = ws.cell(row=row_idx, column=1).value
            if not emp_name or not self.is_paperwork_team(emp_name):
                continue
            
            emp_name = str(emp_name).strip()
            hours = ws.cell(row=row_idx, column=self.current_month + 1).value
            
            try:
                hours = float(hours) if hours else 0
            except:
                hours = 0
            
            self.employees[emp_name]['productivity_hours'] = hours
            self.employees[emp_name]['productivity_status'] = 'Green' if hours >= 8 else 'Red'
        
        wb.close()
        print(f"   ✅ Loaded")
    
    def load_attendance(self):
        print("\n📊 Loading Attendance...")
        file_path = self.folder / "Attendance.xlsx"
        if not file_path.exists():
            return
        
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        
        for row_idx in range(2, ws.max_row + 1):
            emp_name = ws.cell(row=row_idx, column=1).value
            if not emp_name or not self.is_paperwork_team(emp_name):
                continue
            
            emp_name = str(emp_name).strip()
            leaves = ws.cell(row=row_idx, column=self.current_month + 1).value
            
            try:
                leaves = int(leaves) if leaves else 0
            except:
                leaves = 0
            
            attendance_pct = max(0, ((22 - leaves) / 22) * 100)
            self.employees[emp_name]['attendance_leaves'] = leaves
            self.employees[emp_name]['attendance_pct'] = round(attendance_pct, 2)
        
        wb.close()
        print(f"   ✅ Loaded")
    
    def load_pkt(self):
        print("\n📚 Loading PKT...")
        file_path = self.folder / "Process_Knowledge_Test.xlsx"
        if not file_path.exists():
            return
        
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        
        for row_idx in range(2, ws.max_row + 1):
            emp_name = ws.cell(row=row_idx, column=1).value
            if not emp_name or not self.is_paperwork_team(emp_name):
                continue
            
            emp_name = str(emp_name).strip()
            score = ws.cell(row=row_idx, column=self.current_month + 1).value
            
            try:
                score = int(score) if score else 0
            except:
                score = 0
            
            if score >= 90:
                rating = 'Excellent'
            elif score >= 80:
                rating = 'Good'
            else:
                rating = 'Needs Improvement'
            
            self.employees[emp_name]['pkt_score'] = score
            self.employees[emp_name]['pkt_rating'] = rating
        
        wb.close()
        print(f"   ✅ Loaded")
    
    def load_internal_audit(self):
        print("\n🔍 Loading Internal Audit...")
        file_path = self.folder / "Internal_Audit_Scores.xlsx"
        if not file_path.exists():
            return
        
        wb = openpyxl.load_workbook(file_path)
        errors = defaultdict(int)
        
        if 'Internal Audit' in wb.sheetnames:
            ws = wb['Internal Audit']
            for row_idx in range(2, ws.max_row + 1):
                team_member = ws.cell(row=row_idx, column=5).value
                error_count = ws.cell(row=row_idx, column=2).value
                
                if team_member and error_count:
                    if self.is_paperwork_team(team_member):
                        team_member = str(team_member).strip()
                        try:
                            errors[team_member] += int(str(error_count).split()[0])
                        except:
                            pass
        
        for emp_name in errors:
            err = errors[emp_name]
            if err == 0:
                status = 'Green'
            elif err <= 2:
                status = 'Amber'
            else:
                status = 'Red'
            
            self.employees[emp_name]['audit_errors'] = err
            self.employees[emp_name]['audit_status'] = status
        
        wb.close()
        print(f"   ✅ Loaded")
    
    def load_paperwork_clearance(self):
        print("\n📋 Loading Paperwork Clearance SLA...")
        file_path = self.folder / "Paperwork_Clearance_Tracker.xlsx"
        if not file_path.exists():
            return
        
        wb = openpyxl.load_workbook(file_path)
        ws = wb['Clearance Tracker']
        
        clearance_data = defaultdict(lambda: {
            'completed': 0,
            'pending': 0,
            'sla_days': [],
            'sla_met': 0
        })
        
        for row_idx in range(2, min(ws.max_row + 1, 300)):
            coordinator = ws.cell(row=row_idx, column=5).value
            received = ws.cell(row=row_idx, column=1).value
            completed = ws.cell(row=row_idx, column=7).value
            
            if not coordinator or not self.is_paperwork_team(coordinator):
                continue
            
            coordinator = str(coordinator).strip()
            
            if received and completed:
                try:
                    recv_dt = received if isinstance(received, datetime) else datetime.strptime(str(received), '%Y-%m-%d')
                    comp_dt = completed if isinstance(completed, datetime) else datetime.strptime(str(completed), '%Y-%m-%d')
                    sla_days = (comp_dt - recv_dt).days
                    
                    clearance_data[coordinator]['sla_days'].append(sla_days)
                    clearance_data[coordinator]['completed'] += 1
                    if sla_days <= 5:  # Assumed 5-day SLA
                        clearance_data[coordinator]['sla_met'] += 1
                except:
                    pass
        
        for emp_name in clearance_data:
            data = clearance_data[emp_name]
            completed = data['completed']
            sla_days_list = data['sla_days']
            avg_sla = sum(sla_days_list) / len(sla_days_list) if sla_days_list else 0
            sla_compliance = (data['sla_met'] / completed * 100) if completed > 0 else 0
            
            self.employees[emp_name]['paperwork_completed'] = completed
            self.employees[emp_name]['paperwork_avg_sla'] = round(avg_sla, 1)
            self.employees[emp_name]['paperwork_sla_compliance'] = round(sla_compliance, 1)
        
        wb.close()
        print(f"   ✅ Loaded")
    
    def load_paperwork_allocation(self):
        print("\n📍 Loading Paperwork Allocation...")
        file_path = self.folder / "Paperwork_Allocation.xlsx"
        if not file_path.exists():
            return
        
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        
        allocation_data = defaultdict(lambda: {
            'allocated': 0,
            'completed': 0,
            'days_list': []
        })
        
        for row_idx in range(2, min(ws.max_row + 1, 200)):
            allocated_to = ws.cell(row=row_idx, column=6).value
            allocated_date = ws.cell(row=row_idx, column=6).value
            
            if allocated_to and self.is_paperwork_team(allocated_to):
                allocated_to = str(allocated_to).strip()
                allocation_data[allocated_to]['allocated'] += 1
        
        for emp_name in allocation_data:
            data = allocation_data[emp_name]
            self.employees[emp_name]['allocations'] = data['allocated']
            self.employees[emp_name]['allocations_completed'] = data['allocated']
        
        wb.close()
        print(f"   ✅ Loaded")
    
    def load_nh_pending(self):
        print("\n🔔 Loading NH Pending...")
        file_path = self.folder / "New_NH_pending_Tracker.xlsx"
        if not file_path.exists():
            return
        
        wb = openpyxl.load_workbook(file_path)
        ws = wb['2026 NH Pending']
        
        nh_pending = defaultdict(int)
        
        for row_idx in range(2, ws.max_row + 1):
            status = ws.cell(row=row_idx, column=12).value
            processor = ws.cell(row=row_idx, column=7).value
            
            if status and 'Pending' in str(status):
                if processor and self.is_paperwork_team(processor):
                    processor = str(processor).strip()
                    nh_pending[processor] += 1
        
        for emp_name in nh_pending:
            count = nh_pending[emp_name]
            self.employees[emp_name]['nh_pending'] = count
            self.employees[emp_name]['nh_eligibility'] = 'Not Eligible' if count > 0 else 'Eligible'
        
        wb.close()
        print(f"   ✅ Loaded")
    
    def load_client_audit(self):
        print("\n🏢 Loading Client Audit...")
        file_path = self.folder / "Client_System_Audit_Tracker.xlsx"
        if not file_path.exists():
            return
        
        wb = openpyxl.load_workbook(file_path)
        ws = wb['Client System Audit']
        
        audit_pending = defaultdict(int)
        
        for row_idx in range(2, ws.max_row + 1):
            team_member = ws.cell(row=row_idx, column=3).value
            pending = ws.cell(row=row_idx, column=8).value
            
            if team_member and self.is_paperwork_team(team_member):
                team_member = str(team_member).strip()
                try:
                    audit_pending[team_member] += int(pending) if pending else 0
                except:
                    pass
        
        for emp_name in audit_pending:
            count = audit_pending[emp_name]
            self.employees[emp_name]['client_audit_pending'] = count
            self.employees[emp_name]['client_audit_status'] = 'Red' if count > 0 else 'Green'
        
        wb.close()
        print(f"   ✅ Loaded")
    
    def calculate_overall_score(self):
        print("\n🎯 Calculating Scores...")
        
        for emp_name in self.employees:
            emp = self.employees[emp_name]
            
            scores = []
            weights = []
            
            # PKT: 35%
            if 'pkt_score' in emp:
                scores.append(emp['pkt_score'])
                weights.append(35)
            
            # Paperwork SLA: 30%
            if 'paperwork_sla_compliance' in emp:
                scores.append(emp['paperwork_sla_compliance'])
                weights.append(30)
            
            # Productivity: 20%
            if 'productivity_hours' in emp:
                prod = min((emp['productivity_hours'] / 8) * 100, 100)
                scores.append(prod)
                weights.append(20)
            
            # Attendance: 15%
            if 'attendance_pct' in emp:
                scores.append(emp['attendance_pct'])
                weights.append(15)
            
            if scores:
                overall = sum(s * w for s, w in zip(scores, weights)) / sum(weights) if weights else 0
                emp['overall_score'] = round(overall, 2)
                emp['overall_rating'] = 'Excellent' if overall >= 90 else 'Good' if overall >= 80 else 'Needs Improvement'
                
                nh = emp.get('nh_pending', 0)
                audit = emp.get('client_audit_pending', 0)
                emp['incentive_eligible'] = 'Not Eligible' if (nh > 0 or audit > 0) else 'Eligible'
    
    def generate_scorecard(self):
        print("\n📊 Generating Scorecard...")
        
        scorecard = []
        for emp_name in sorted(self.employees.keys()):
            emp = self.employees[emp_name]
            
            scorecard.append({
                'Employee Name': emp_name,
                'Productivity Hours': emp.get('productivity_hours', '-'),
                'Productivity Status': emp.get('productivity_status', '-'),
                'NH Pending': emp.get('nh_pending', 0),
                'NH Eligibility': emp.get('nh_eligibility', '-'),
                'Audit Errors': emp.get('audit_errors', 0),
                'Audit Status': emp.get('audit_status', '-'),
                'Paperwork Completed': emp.get('paperwork_completed', 0),
                'Paperwork SLA Avg': emp.get('paperwork_avg_sla', '-'),
                'SLA Compliance %': emp.get('paperwork_sla_compliance', '-'),
                'Allocations': emp.get('allocations', 0),
                'Allocations Completed': emp.get('allocations_completed', 0),
                'PKT Score': emp.get('pkt_score', '-'),
                'PKT Rating': emp.get('pkt_rating', '-'),
                'Attendance Leaves': emp.get('attendance_leaves', '-'),
                'Attendance %': emp.get('attendance_pct', '-'),
                'Client Audit Pending': emp.get('client_audit_pending', 0),
                'Client Audit Status': emp.get('client_audit_status', '-'),
                'Overall Score': emp.get('overall_score', '-'),
                'Overall Rating': emp.get('overall_rating', '-'),
                'Incentive Eligible': emp.get('incentive_eligible', '-'),
                'Remarks': 'On Track' if emp.get('incentive_eligible') == 'Eligible' else 'Action Required'
            })
        
        return scorecard
    
    def run(self):
        print("\n" + "="*80)
        print("PAPERWORK CLEARANCE TEAM MONTHLY SCORECARD GENERATOR")
        print("="*80)
        
        self.load_production()
        self.load_attendance()
        self.load_pkt()
        self.load_internal_audit()
        self.load_paperwork_clearance()
        self.load_paperwork_allocation()
        self.load_nh_pending()
        self.load_client_audit()
        
        self.calculate_overall_score()
        scorecard = self.generate_scorecard()
        
        print("\n" + "="*80)
        print(f"✅ SCORECARD GENERATED FOR {len(scorecard)} PAPERWORK CLEARANCE TEAM MEMBERS")
        print("="*80)
        
        return scorecard, self.employees

if __name__ == "__main__":
    folder_path = input("\n📁 Enter folder path: ").strip().strip('"')
    
    generator = PaperworkClearanceScorecard(folder_path)
    scorecard, employees = generator.run()
    
    with open('/mnt/user-data/outputs/paperwork_clearance_scorecard.json', 'w') as f:
        json.dump(scorecard, f, indent=2, default=str)
    
    print(f"\n✅ Saved: paperwork_clearance_scorecard.json")
    
    if scorecard:
        print(f"\n{'Employee':<25} {'Score':<8} {'Rating':<15} {'SLA %':<10} {'Incentive':<15}")
        print("-" * 75)
        for row in scorecard:
            print(f"{row['Employee Name']:<25} {str(row['Overall Score']):<8} {str(row['Overall Rating']):<15} {str(row['SLA Compliance %']):<10} {str(row['Incentive Eligible']):<15}")
