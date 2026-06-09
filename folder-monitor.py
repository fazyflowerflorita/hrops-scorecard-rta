#!/usr/bin/env python3
"""
HR Operations Scorecard - Local Folder Monitor & Firebase Sync
Monitors a local folder (synced with SharePoint) and automatically syncs data to Firebase
"""

import os
import sys
import json
import time
from pathlib import Path
import openpyxl
from openpyxl.utils import get_column_letter
import firebase_admin
from firebase_admin import credentials
from firebase_admin import db
from datetime import datetime
import hashlib

# Firebase Configuration
FIREBASE_CONFIG = {
    "apiKey": "AIzaSyD_ToRRVy18YjH_V0M67DZ9Pd6LcSqDQqg",
    "authDomain": "hrops-scorecard---rta.firebaseapp.com",
    "projectId": "hrops-scorecard---rta",
    "databaseURL": "https://hrops-scorecard---rta.firebaseio.com",
    "storageBucket": "hrops-scorecard---rta.firebasestorage.app",
    "messagingSenderId": "127519107306",
    "appId": "1:127519107306:web:61683e0eb03874b0a94f93",
    "measurementId": "G-K5H5SV6QQE"
}

class FirebaseInitializer:
    """Handle Firebase initialization"""
    
    @staticmethod
    def initialize():
        """Initialize Firebase with service account"""
        try:
            # Create a service account JSON file (user needs to download from Firebase Console)
            service_account_path = "firebase-service-account.json"
            
            if not os.path.exists(service_account_path):
                print("❌ firebase-service-account.json not found!")
                print("\n📝 To fix this:")
                print("1. Go to: https://console.firebase.google.com/")
                print("2. Select: HROps Scorecard - RTA project")
                print("3. Click: Settings (gear icon) → Service Accounts")
                print("4. Click: Generate New Private Key")
                print("5. Save as: firebase-service-account.json in this folder")
                sys.exit(1)
            
            cred = credentials.Certificate(service_account_path)
            app = firebase_admin.initialize_app(cred, {
                'databaseURL': FIREBASE_CONFIG['databaseURL']
            })
            print("✅ Firebase initialized successfully!")
            return db.reference()
        except Exception as e:
            print(f"❌ Firebase initialization error: {e}")
            sys.exit(1)

class ExcelFileReader:
    """Read and parse Excel files"""
    
    @staticmethod
    def read_excel_file(file_path):
        """Read Excel file and return data as list of dicts"""
        try:
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            
            # Get headers from first row
            headers = []
            for cell in ws[1]:
                if cell.value:
                    headers.append(str(cell.value).strip())
            
            # Read data rows
            data = []
            for row in ws.iter_rows(min_row=2, values_only=False):
                row_data = {}
                for idx, cell in enumerate(row):
                    if idx < len(headers):
                        value = cell.value
                        # Convert None to empty string
                        row_data[headers[idx]] = value if value is not None else ""
                
                # Only add if row has data
                if any(row_data.values()):
                    data.append(row_data)
            
            wb.close()
            return data, headers
        except Exception as e:
            print(f"❌ Error reading {file_path}: {e}")
            return [], []

class ScorecardDataProcessor:
    """Process and sync scorecard data to Firebase"""
    
    TEAMS = [
        'Compliance',
        'Final Clearance Team',
        'HR Operations',
        'Internal Audit Team',
        'Paperwork Audit',
        'MSD Data Management',
        'Requisitions and Submissions'
    ]
    
    # Column mappings
    EMPLOYEE_COLUMNS = {
        'Employee. ID': ['empId', 'employee_id', 'id'],
        'Names': ['name', 'employee_name'],
        'Department': ['dept', 'department'],
        'Team': ['team'],
        'India Reporting Manager': ['manager', 'reporting_manager']
    }
    
    PERFORMANCE_COLUMNS = {
        'empId': ['employee_id', 'id'],
        'overall_score': ['overall score', 'overall_score', 'score'],
        'quality_score': ['quality score', 'quality_score'],
        'attendance': ['attendance'],
        'process_knowledge': ['process knowledge', 'process_knowledge', 'knowledge'],
        'target_achievement': ['target achievement', 'target_achievement', 'target']
    }
    
    @staticmethod
    def normalize_column_name(actual_name, expected_columns):
        """Normalize column name to match expected format"""
        if not actual_name:
            return None
        
        actual_lower = str(actual_name).lower().strip()
        
        for expected, alternatives in expected_columns.items():
            if actual_lower == expected.lower():
                return expected
            for alt in alternatives:
                if actual_lower == alt.lower():
                    return expected
        
        return None
    
    @staticmethod
    def process_employee_data(data, headers):
        """Process employee master data"""
        employees = {}
        
        for row in data:
            emp_data = {}
            emp_id = None
            
            for key, value in row.items():
                normalized = ScorecardDataProcessor.normalize_column_name(
                    key, 
                    ScorecardDataProcessor.EMPLOYEE_COLUMNS
                )
                
                if normalized:
                    emp_data[normalized.lower().replace(' ', '_')] = str(value).strip() if value else ""
                    
                    if normalized == 'Employee. ID':
                        emp_id = str(value).strip() if value else None
            
            if emp_id and emp_data.get('names'):
                employees[emp_id] = {
                    'empId': emp_id,
                    'name': emp_data.get('names', ''),
                    'dept': emp_data.get('department', ''),
                    'team': emp_data.get('team', ''),
                    'manager': emp_data.get('india_reporting_manager', '')
                }
        
        return employees
    
    @staticmethod
    def process_performance_data(data):
        """Process performance metrics data"""
        performance = {}
        
        for row in data:
            perf_data = {}
            emp_id = None
            
            for key, value in row.items():
                normalized = ScorecardDataProcessor.normalize_column_name(
                    key,
                    ScorecardDataProcessor.PERFORMANCE_COLUMNS
                )
                
                if normalized:
                    try:
                        num_value = float(value) if value else 0
                    except (ValueError, TypeError):
                        num_value = 0
                    
                    perf_data[normalized] = num_value
                    
                    if normalized == 'empId':
                        emp_id = str(value).strip() if value else None
            
            if emp_id:
                performance[emp_id] = {
                    'empId': emp_id,
                    'overall_score': perf_data.get('overall_score', 0),
                    'quality_score': perf_data.get('quality_score', 0),
                    'attendance': perf_data.get('attendance', 0),
                    'process_knowledge': perf_data.get('process_knowledge', 0),
                    'target_achievement': perf_data.get('target_achievement', 0)
                }
        
        return performance

class FolderMonitor:
    """Monitor local folder for Excel file changes"""
    
    def __init__(self, folder_path, db_ref):
        self.folder_path = Path(folder_path)
        self.db_ref = db_ref
        self.file_hashes = {}
        self.last_sync = None
        
        if not self.folder_path.exists():
            print(f"❌ Folder not found: {folder_path}")
            sys.exit(1)
        
        print(f"✅ Monitoring folder: {self.folder_path}")
    
    def get_file_hash(self, file_path):
        """Get hash of file to detect changes"""
        try:
            with open(file_path, 'rb') as f:
                return hashlib.md5(f.read()).hexdigest()
        except:
            return None
    
    def find_excel_files(self):
        """Find all Excel files in folder"""
        excel_files = []
        for file_path in self.folder_path.rglob('*.xlsx'):
            if not file_path.name.startswith('~$'):  # Skip temp files
                excel_files.append(file_path)
        for file_path in self.folder_path.rglob('*.xls'):
            if not file_path.name.startswith('~$'):
                excel_files.append(file_path)
        return excel_files
    
    def process_files(self):
        """Check for file changes and sync to Firebase"""
        excel_files = self.find_excel_files()
        
        if not excel_files:
            print(f"⏳ Waiting for Excel files in {self.folder_path.name}...")
            return False
        
        print(f"\n📁 Found {len(excel_files)} Excel file(s):")
        
        all_employees = {}
        all_performance = {}
        file_changed = False
        
        for file_path in excel_files:
            print(f"  📄 {file_path.name}")
            
            current_hash = self.get_file_hash(file_path)
            file_key = str(file_path)
            
            # Check if file changed
            if file_key not in self.file_hashes or self.file_hashes[file_key] != current_hash:
                self.file_hashes[file_key] = current_hash
                file_changed = True
                
                # Read file
                data, headers = ExcelFileReader.read_excel_file(file_path)
                
                if not data:
                    continue
                
                # Detect file type by content
                if any('Employee. ID' in str(h) for h in headers) or any('empId' in str(h) for h in headers):
                    # Might be employee master or performance data
                    if any('overall_score' in str(h).lower() or 'Overall Score' in str(h) for h in headers):
                        # Performance data
                        perf = ScorecardDataProcessor.process_performance_data(data)
                        all_performance.update(perf)
                        print(f"    ✅ Synced {len(perf)} performance records")
                    else:
                        # Employee master
                        emps = ScorecardDataProcessor.process_employee_data(data, headers)
                        all_employees.update(emps)
                        print(f"    ✅ Synced {len(emps)} employees")
        
        # Sync to Firebase
        if file_changed and (all_employees or all_performance):
            self.sync_to_firebase(all_employees, all_performance)
            self.last_sync = datetime.now()
            return True
        
        return False
    
    def sync_to_firebase(self, employees, performance):
        """Sync data to Firebase"""
        try:
            if employees:
                self.db_ref.child('employee-master').set(employees)
                print(f"\n☁️ Synced {len(employees)} employees to Firebase")
            
            if performance:
                self.db_ref.child('performance-data').set(performance)
                print(f"☁️ Synced {len(performance)} performance records to Firebase")
            
            print(f"✅ Last sync: {self.last_sync.strftime('%Y-%m-%d %H:%M:%S')}")
        except Exception as e:
            print(f"❌ Firebase sync error: {e}")
    
    def start_monitoring(self, interval=5):
        """Start monitoring folder for changes"""
        print(f"\n🚀 Starting folder monitor (checking every {interval} seconds)")
        print("Press Ctrl+C to stop...\n")
        
        try:
            while True:
                try:
                    self.process_files()
                except Exception as e:
                    print(f"❌ Error: {e}")
                
                time.sleep(interval)
        except KeyboardInterrupt:
            print("\n\n✅ Monitor stopped.")
            sys.exit(0)

def main():
    """Main function"""
    print("╔════════════════════════════════════════════════════════════════╗")
    print("║   HR OPERATIONS SCORECARD - Local Folder Monitor & Sync       ║")
    print("╚════════════════════════════════════════════════════════════════╝\n")
    
    # Get folder path from user
    print("Enter the local folder path (synced with SharePoint):")
    print("Example: C:\\Users\\FazyFlowerFlorita\\Pride Technologies\\RTA - 2026\n")
    
    folder_path = input("📁 Folder path: ").strip().strip('"')
    
    if not folder_path:
        print("❌ No folder path provided!")
        sys.exit(1)
    
    # Initialize Firebase
    print("\n🔄 Initializing Firebase...")
    db_ref = FirebaseInitializer.initialize()
    
    # Start monitoring
    print(f"\n🔄 Initializing folder monitor...")
    monitor = FolderMonitor(folder_path, db_ref)
    monitor.start_monitoring(interval=5)

if __name__ == "__main__":
    main()
