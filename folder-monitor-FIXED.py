#!/usr/bin/env python3
"""
HR Operations Scorecard - Local Folder Monitor & Firebase Sync (FIXED)
✅ Uses correct Firebase Admin SDK syntax
"""

import os
import sys
import time
from pathlib import Path
import openpyxl
import firebase_admin
from firebase_admin import credentials, db
from datetime import datetime
import hashlib

# Firebase Configuration
FIREBASE_CONFIG = {
    "projectId": "hrops-scorecard---rta",
    "databaseURL": "https://hrops-scorecard---rta.firebaseio.com"
}

class FirebaseInitializer:
    """Handle Firebase initialization"""
    
    @staticmethod
    def initialize():
        """Initialize Firebase with service account"""
        try:
            service_account_path = "firebase-service-account.json"
            
            if not os.path.exists(service_account_path):
                print("❌ firebase-service-account.json not found!")
                print("\n📝 To fix this:")
                print("1. Make sure the file is in the SAME folder as this script")
                print("2. File must be named EXACTLY: firebase-service-account.json")
                sys.exit(1)
            
            # Delete any existing app to avoid conflicts
            if firebase_admin._apps:
                firebase_admin.delete_app(firebase_admin.get_app())
            
            cred = credentials.Certificate(service_account_path)
            app = firebase_admin.initialize_app(cred, {
                'databaseURL': FIREBASE_CONFIG['databaseURL']
            })
            print("✅ Firebase initialized successfully!")
            return True
        except Exception as e:
            print(f"❌ Firebase initialization error: {e}")
            return False

class ExcelFileReader:
    """Read and parse Excel files"""
    
    @staticmethod
    def read_excel_file(file_path):
        """Read Excel file and return data as list of dicts"""
        try:
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            
            # Get headers from first row
            headers = []
            for cell in ws[1]:
                if cell.value:
                    headers.append(str(cell.value).strip())
            
            # Read data rows
            data = []
            for row in ws.iter_rows(min_row=2, values_only=False):
                row_data = {}
                for idx, cell in enumerate(row):
                    if idx < len(headers):
                        value = cell.value
                        row_data[headers[idx]] = value if value is not None else ""
                
                # Only add if row has data
                if any(row_data.values()):
                    data.append(row_data)
            
            wb.close()
            return data, headers
        except Exception as e:
            print(f"❌ Error reading {file_path}: {e}")
            return [], []

class ScorecardDataProcessor:
    """Process and sync scorecard data to Firebase"""
    
    @staticmethod
    def normalize_column_name(actual_name, expected_columns):
        """Normalize column name to match expected format"""
        if not actual_name:
            return None
        
        actual_lower = str(actual_name).lower().strip()
        
        for expected, alternatives in expected_columns.items():
            if actual_lower == expected.lower():
                return expected
            for alt in alternatives:
                if actual_lower == alt.lower():
                    return expected
        
        return None
    
    @staticmethod
    def process_employee_data(data, headers):
        """Process employee master data"""
        employees = {}
        
        EMPLOYEE_COLUMNS = {
            'Employee. ID': ['empId', 'employee_id', 'id', 'emp id'],
            'Names': ['name', 'employee_name', 'employee name'],
            'Department': ['dept', 'department', 'deptartment'],
            'Team': ['team'],
            'India Reporting Manager': ['manager', 'reporting_manager', 'reporting manager']
        }
        
        for row in data:
            emp_data = {}
            emp_id = None
            
            for key, value in row.items():
                normalized = ScorecardDataProcessor.normalize_column_name(
                    key, 
                    EMPLOYEE_COLUMNS
                )
                
                if normalized:
                    emp_data[normalized.lower().replace(' ', '_')] = str(value).strip() if value else ""
                    
                    if normalized == 'Employee. ID':
                        emp_id = str(value).strip() if value else None
            
            if emp_id:
                employees[emp_id] = {
                    'empId': emp_id,
                    'name': emp_data.get('names', ''),
                    'dept': emp_data.get('department', ''),
                    'team': emp_data.get('team', ''),
                    'manager': emp_data.get('india_reporting_manager', '')
                }
        
        return employees
    
    @staticmethod
    def process_performance_data(data):
        """Process performance metrics data"""
        performance = {}
        
        PERFORMANCE_COLUMNS = {
            'empId': ['employee_id', 'id', 'emp id'],
            'overall_score': ['overall score', 'overall_score', 'score'],
            'quality_score': ['quality score', 'quality_score'],
            'attendance': ['attendance'],
            'process_knowledge': ['process knowledge', 'process_knowledge', 'knowledge'],
            'target_achievement': ['target achievement', 'target_achievement', 'target']
        }
        
        for row in data:
            perf_data = {}
            emp_id = None
            
            for key, value in row.items():
                normalized = ScorecardDataProcessor.normalize_column_name(
                    key,
                    PERFORMANCE_COLUMNS
                )
                
                if normalized:
                    try:
                        num_value = float(value) if value else 0
                    except (ValueError, TypeError):
                        num_value = 0
                    
                    perf_data[normalized] = num_value
                    
                    if normalized == 'empId':
                        emp_id = str(value).strip() if value else None
            
            if emp_id:
                performance[emp_id] = {
                    'empId': emp_id,
                    'overall_score': perf_data.get('overall_score', 0),
                    'quality_score': perf_data.get('quality_score', 0),
                    'attendance': perf_data.get('attendance', 0),
                    'process_knowledge': perf_data.get('process_knowledge', 0),
                    'target_achievement': perf_data.get('target_achievement', 0)
                }
        
        return performance

class FolderMonitor:
    """Monitor local folder for Excel file changes"""
    
    def __init__(self, folder_path):
        self.folder_path = Path(folder_path)
        self.file_hashes = {}
        self.last_sync = None
        
        if not self.folder_path.exists():
            print(f"❌ Folder not found: {folder_path}")
            sys.exit(1)
        
        print(f"✅ Monitoring folder: {self.folder_path}")
    
    def get_file_hash(self, file_path):
        """Get hash of file to detect changes"""
        try:
            with open(file_path, 'rb') as f:
                return hashlib.md5(f.read()).hexdigest()
        except:
            return None
    
    def find_excel_files(self):
        """Find all Excel files in folder"""
        excel_files = []
        for file_path in self.folder_path.rglob('*.xlsx'):
            if not file_path.name.startswith('~$'):
                excel_files.append(file_path)
        for file_path in self.folder_path.rglob('*.xls'):
            if not file_path.name.startswith('~$'):
                excel_files.append(file_path)
        return excel_files
    
    def process_files(self):
        """Check for file changes and sync to Firebase"""
        excel_files = self.find_excel_files()
        
        if not excel_files:
            return False
        
        all_employees = {}
        all_performance = {}
        file_changed = False
        
        for file_path in excel_files:
            current_hash = self.get_file_hash(file_path)
            file_key = str(file_path)
            
            # Check if file changed
            if file_key not in self.file_hashes or self.file_hashes[file_key] != current_hash:
                self.file_hashes[file_key] = current_hash
                file_changed = True
                
                # Read file
                data, headers = ExcelFileReader.read_excel_file(file_path)
                
                if not data:
                    continue
                
                print(f"\n📊 Processing: {file_path.name}")
                
                # Detect file type by content
                if any('Employee' in str(h) or 'empId' in str(h) for h in headers):
                    if any('score' in str(h).lower() for h in headers):
                        # Performance data
                        perf = ScorecardDataProcessor.process_performance_data(data)
                        all_performance.update(perf)
                        print(f"   ✅ Parsed {len(perf)} performance records")
                    else:
                        # Employee master
                        emps = ScorecardDataProcessor.process_employee_data(data, headers)
                        all_employees.update(emps)
                        print(f"   ✅ Parsed {len(emps)} employees")
        
        # Sync to Firebase
        if file_changed and (all_employees or all_performance):
            self.sync_to_firebase(all_employees, all_performance)
            self.last_sync = datetime.now()
            return True
        
        return False
    
    def sync_to_firebase(self, employees, performance):
        """Sync data to Firebase using correct Admin SDK syntax"""
        try:
            # ✅ CORRECT SYNTAX for Firebase Admin SDK:
            # Use db.reference('path').set(data)
            
            if employees:
                db.reference('employee-master').set(employees)
                print(f"\n☁️ Synced {len(employees)} employees to Firebase ✅")
            
            if performance:
                db.reference('performance-data').set(performance)
                print(f"☁️ Synced {len(performance)} performance records to Firebase ✅")
            
            print(f"✅ Last sync: {self.last_sync.strftime('%Y-%m-%d %H:%M:%S')}")
        except Exception as e:
            print(f"❌ Firebase sync error: {e}")
            print(f"\nDEBUG INFO:")
            print(f"  Database URL: {FIREBASE_CONFIG['databaseURL']}")
            print(f"  Error type: {type(e).__name__}")
    
    def start_monitoring(self, interval=10):
        """Start monitoring folder for changes"""
        print(f"\n🚀 Starting folder monitor (checking every {interval} seconds)")
        print("Press Ctrl+C to stop...\n")
        
        try:
            check_count = 0
            while True:
                check_count += 1
                excel_files = self.find_excel_files()
                
                if excel_files:
                    print(f"\n[{check_count}] 📁 Found {len(excel_files)} Excel file(s)")
                    self.process_files()
                else:
                    print(f"[{check_count}] ⏳ No Excel files found yet...")
                
                time.sleep(interval)
        except KeyboardInterrupt:
            print("\n\n✅ Monitor stopped.")
            sys.exit(0)

def main():
    """Main function"""
    print("\n╔════════════════════════════════════════════════════════════════╗")
    print("║   HR OPERATIONS SCORECARD - Folder Monitor & Sync (FIXED)   ║")
    print("╚════════════════════════════════════════════════════════════════╝\n")
    
    # Initialize Firebase first
    print("🔄 Initializing Firebase...")
    if not FirebaseInitializer.initialize():
        print("\n❌ Firebase initialization failed!")
        sys.exit(1)
    
    # Get folder path from user
    print("\nEnter your local folder path (where Excel files are):")
    print("Example: C:\\Users\\FazyFlowerFlorita\\Pride Technologies\\RTA - 2026\n")
    
    folder_path = input("📁 Folder path: ").strip().strip('"')
    
    if not folder_path:
        print("❌ No folder path provided!")
        sys.exit(1)
    
    # Start monitoring
    print(f"\n🔄 Initializing folder monitor...")
    monitor = FolderMonitor(folder_path)
    monitor.start_monitoring(interval=10)

if __name__ == "__main__":
    main()
